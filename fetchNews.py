from requests import Session
import requests
import json
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from bs4 import BeautifulSoup

#########CryptoPanic############################################################################################################################

wb = Workbook()
sheet = wb.active
today = date.today()
i = 1
session = Session()
urls = ['https://cryptopanic.com/api/v1/posts/?auth_token=f8d35df861317a27f9e520363e61f2e083fea36e&public=true&filter=important',
              'https://cryptopanic.com/api/v1/posts/?auth_token=f8d35df861317a27f9e520363e61f2e083fea36e&public=true&filter=hot',
              'https://cryptopanic.com/api/v1/posts/?auth_token=f8d35df861317a27f9e520363e61f2e083fea36e&public=true']


for public_url in urls:
    public_response = session.get(public_url)
    pubic_data = json.loads(public_response.text)["results"]

    for public_new in pubic_data:
        ##TODO:create content
        content = []
        title = public_new["title"]
        publish_date = public_new["created_at"]
        url = public_new["url"]

        content.append(title)
        #TODO: adjust datetime format
        # publish_date = datetime.strptime(publish_date, "%Y-%m-%dT%H:%M:%SZ") + timedelta(hours=8)

        print(f"{i}: {title}\n{url}\n")


        # Add a hyperlink
        sheet.cell(row=i, column=1).value = '=HYPERLINK("{}", "{}")'.format(url, title)
        wb.save(f"files/News_{today}.xlsx")

        i += 1

#######The Block Latest###########################################################################################################################################

##4 pages of news = 40
url_list = ['https://www.theblock.co/latest']

for n in range(10, 40, 10):
    url_list.append(f'https://www.theblock.co/latest?start={n}')

for url in url_list:
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    result_list = soup.findAll('article')
    ##headline
    for k in range(0, len(result_list)):
        result = result_list[k]
        head_msg = result.findNext('span')
        head = head_msg['ga-event-label'][12:]

        ##link
        link_msg = result.findNext('a')
        link = f"https://www.theblock.co{str(link_msg['href'])}"

        print(f"{i}: {head}\n{link}\n")
        sheet.cell(row=i, column=1).value = '=HYPERLINK("{}", "{}")'.format(link, head)
        wb.save(f"files/News_{today}.xlsx")
        i += 1


print(f"Done! Please find excel file News_{today}")
